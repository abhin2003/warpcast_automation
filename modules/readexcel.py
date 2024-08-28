import openpyxl
import re

def get_id():
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook("Google Search Result_Bitcoin.xlsx")

    # Define variable to read sheet
    dataframe1 = dataframe.active

    for row in range(3, dataframe1.max_row):
        
        identiy = dataframe1.cell(row=row, column=2).value
        
        start = identiy.find('(')
        end = identiy.find(')')
        result = identiy[start+2:end]
        print(result)


get_id()